"""
Build deployment-log.xlsx from the release-generator's out/deployment-rows.json.

Rebuilds the log fresh from the current release each run, so editing release.json
and re-running keeps the Excel log in lockstep with the email / Teams / CR outputs.

Usage: python release/scripts/build-xlsx.py
"""

import json
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

HERE = os.path.dirname(os.path.abspath(__file__))
RELEASE_DIR = os.path.dirname(HERE)
ROWS_JSON = os.path.join(RELEASE_DIR, "out", "deployment-rows.json")
OUT_XLSX = os.path.join(RELEASE_DIR, "deployment-log.xlsx")

with open(ROWS_JSON, "r", encoding="utf-8") as f:
    data = json.load(f)
rows = data["rows"]

COLUMNS = [
    ("Release", "release", 16),
    ("Story", "story", 10),
    ("Summary", "summary", 44),
    ("Feature", "feature", 28),
    ("Environment", "environment", 18),
    ("Deploy Date", "deployDate", 13),
    ("Window", "deployWindow", 14),
    ("Deployed By", "deployedBy", 14),
    ("Approver", "approver", 16),
    ("Change Request", "changeRequest", 16),
    ("Risk", "riskLevel", 9),
    ("Points", "storyPoints", 8),
    ("Status", "status", 12),
]

wb = Workbook()
ws = wb.active
ws.title = "Deployment Log"

header_fill = PatternFill("solid", fgColor="0F766E")
header_font = Font(color="FFFFFF", bold=True, size=11)
thin = Side(style="thin", color="D0D7DE")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
risk_fill = {
    "Low": PatternFill("solid", fgColor="DCFCE7"),
    "Medium": PatternFill("solid", fgColor="FEF9C3"),
    "High": PatternFill("solid", fgColor="FEE2E2"),
}

# header
for c, (label, _key, width) in enumerate(COLUMNS, start=1):
    cell = ws.cell(row=1, column=c, value=label)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(vertical="center", horizontal="left")
    cell.border = border
    ws.column_dimensions[cell.column_letter].width = width
ws.row_dimensions[1].height = 22
ws.freeze_panes = "A2"

# rows
for r, row in enumerate(rows, start=2):
    for c, (_label, key, _w) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=r, column=c, value=row.get(key, ""))
        cell.border = border
        cell.alignment = Alignment(vertical="center", horizontal="left", wrap_text=(key == "summary"))
        if key == "riskLevel" and row.get(key) in risk_fill:
            cell.fill = risk_fill[row[key]]
        if key == "story":
            cell.font = Font(bold=True)

# a light banner row note under the table
note_row = len(rows) + 3
ws.cell(row=note_row, column=1,
        value=f"Generated from release.json — {data['release']} — {len(rows)} stories promoted to production (L3).").font = Font(
    italic=True, color="64748B", size=10)

wb.save(OUT_XLSX)
print(f"Wrote: {os.path.relpath(OUT_XLSX, RELEASE_DIR)}  ({len(rows)} rows)")
